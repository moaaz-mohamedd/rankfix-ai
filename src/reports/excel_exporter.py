from io import BytesIO # this is used to create an in memory file not in the disk and return it as bytes for streamlit to download

import pandas as pd

## openpyxl is used to create and style(like align and wrap_text and font) the excel file in memory
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter # convert column index to letter for styling and adjusting width


def create_excel_report(df: pd.DataFrame) -> bytes:
    """
    Create an Excel report from the SEO audit dataframe.

    This function does not save the file on disk.
    It creates the Excel file in memory so Streamlit can download it directly.
    """

    output = BytesIO() ## Create an in-memory bytes buffer to hold the Excel file data

    with pd.ExcelWriter(output, engine="openpyxl") as writer: # create an excel writer in the in memory buffer we made above
        sheet_name = "SEO Audit"

        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name
        )
        # workbook is a refrence to the entire excel file we made, and worksheet is a refrence to the specific sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Style header row
        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze first row
        worksheet.freeze_panes = "A2"

        # Auto-adjust column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                value_length = len(str(value))

                if value_length > max_length:
                    max_length = value_length

            adjusted_width = min(max(max_length + 2, 12), 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Wrap text for long columns
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

    output.seek(0)

    return output.getvalue()